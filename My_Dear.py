#!/usr/bin/python
#coding:utf-8
from PyQt4.QtCore import *  
from PyQt4.QtGui import *  
import openpyxl
import warnings
import sys
import datetime

# 給妹子写的xlsx处理软件
QTextCodec.setCodecForTr(QTextCodec.codecForName("utf8"))  

warnings.filterwarnings("ignore")
  
class XlsProcess(QDialog):  
    def __init__(self, parent = None):  
        super(XlsProcess, self).__init__(parent)  
        self.setWindowTitle(self.tr("Xlsx文件处理工具"))  
  
        mainSplitter = QSplitter(Qt.Horizontal)  
        mainSplitter.setOpaqueResize(True)  
   
        listWidget = QListWidget(mainSplitter)  
        listWidget.insertItem(0, self.tr("考勤文件处理"))  
        listWidget.insertItem(1, self.tr("xlsx文件对比"))   
  
        frame = QFrame(mainSplitter)  
        stack = QStackedWidget()  

        mainSplitter.setStretchFactor(0, 30)
        mainSplitter.setStretchFactor(1, 70)
          
        registration = Registration()  
        comparison = Comparison()  
        stack.addWidget(registration)  
        stack.addWidget(comparison)  
          
        mainLayout = QVBoxLayout(frame)  
        mainLayout.addWidget(stack)  

        info = QLabel(self.tr("Copyright By 伦大锤 - 为心爱的比比开发 "))
        closeButton = QPushButton(self.tr("关闭"))
        bottomLayout = QHBoxLayout()
        bottomLayout.addStretch(1)  
        bottomLayout.addWidget(info)
        bottomLayout.addWidget(closeButton)

        self.connect(listWidget, SIGNAL("currentRowChanged(int)"), stack, SLOT("setCurrentIndex(int)"))  
        self.connect(closeButton, SIGNAL("clicked()"), self, SLOT("close()"))
        mainLayout.addLayout(bottomLayout)
  
        layout = QHBoxLayout(self)  
        layout.addWidget(mainSplitter)      
        self.setLayout(layout)

class Registration(QWidget):  
    def __init__(self, parent = None):  
        super(Registration, self).__init__(parent)  

        filePushButton = QPushButton(self.tr("选择考勤文件"))    
        self.fileLineText = QLineEdit()

        infoLabel = QLabel(self.tr("包含人员及日期"))
        self.infoText = QTextEdit()

        otherLabel = QLabel(self.tr("该段时间内可能的人员异动"))
        self.otherLineText = QLineEdit()

        self.resultTable = QTableWidget()
        self.resultTable.setColumnCount(5)
        self.resultTable.setRowCount(0)
        self.resultTable.verticalHeader().setVisible(False)
        self.resultTable.setEditTriggers(QAbstractItemView.NoEditTriggers)

        self.resultTable.setHorizontalHeaderLabels([self.tr('日期'), self.tr('未来'), self.tr('迟到'), self.tr('早退'), self.tr('消失')]) 
        
        mainLayout = QGridLayout(self)
        mainLayout.addWidget(filePushButton, 0, 0)
        mainLayout.addWidget(self.fileLineText, 0, 1)
        mainLayout.addWidget(infoLabel, 1, 0)
        mainLayout.addWidget(self.infoText, 1, 1)
        mainLayout.addWidget(self.resultTable, 2, 0, 1, 2)
        mainLayout.addWidget(otherLabel, 3, 0)
        mainLayout.addWidget(self.otherLineText, 3, 1)

        self.connect(filePushButton, SIGNAL("clicked()"), self.openFile)

    def openFile(self):
        s = QFileDialog.getOpenFileName(self, "Please Choose A File", ".", "Xlsx files(*.xlsx)")
        self.fileLineText.setText(s)  
        s = str(s.toUtf8())
        s = s.split('/')[-1].decode("utf8")
        work_book = openpyxl.load_workbook(s)
        sheet = work_book.get_sheet_by_name(work_book.get_sheet_names()[0])

        R = sheet.get_highest_row()
        C = sheet.get_highest_column()

        persons = {}
        days = {}

        for r in xrange(2, R):
                key = sheet.cell(row = r, column = 4).value 
                if key != None and (not days.has_key(key)):
                        days[key] = 1

                key = sheet.cell(row = r, column = 3).value
                if key != None and (not persons.has_key(key)):
                        persons[key] = {}

        days = sorted(days.keys())

        temp = self.tr("-------- 相关人员(共" + str(len(persons)) + "人) --------\n")
        for person in persons.keys():
                temp = temp + person + '\n'
        temp = temp + self.tr("\n-------- 包含日期(共" + str(len(days)) + "天) --------\n")
        for day in days:
                temp = temp + day + '\n'
        self.infoText.setText(temp[:-1]) 

        Col = 0
        for key in days:
                online = []
                offline = []
                absent = []
                disappear = []

                for person in persons.values():
                        person['online'] = 0
                        person['offline'] = 0
                        person['absent'] = 1
                        person['disappear'] = 1

                for r in xrange(2,R):
                        if sheet.cell(row=r, column=4).value == key:
                                regTime = sheet.cell(row=r, column=5).value
                                name = sheet.cell(row=r, column=3).value
                                if regTime != None:
                                        hour = int(regTime.split(':')[0])
                                        minute = int(regTime.split(':')[1])
                                        if (hour < 9) or (hour == 9 and minute <= 30):
                                                persons[name]['online'] = 1
                                        elif hour < 12:
                                                persons[name]['absent'] = 0
                                        elif (hour > 18) or (hour == 18 and minute >= 30):
                                                persons[name]['offline'] = 1
                                        elif hour >= 12:
                                                persons[name]['disappear'] = 0

                self.resultTable.setRowCount(len(days))  
                newItem = QTableWidgetItem(self.tr(key))
                newItem.setTextAlignment(Qt.AlignCenter)
                self.resultTable.setItem(Col, 0, newItem)

                # 未来
                temp = ''
                for k, v in persons.items():
                        if v['online'] == 0 and v['absent'] == 1:
                                temp = temp + k + '\n'
                newItem = QTableWidgetItem(self.tr(temp[:-1]))
                newItem.setTextAlignment(Qt.AlignCenter)
                self.resultTable.setItem(Col, 1, newItem)

                # 迟到
                temp = ''
                for k, v in persons.items():
                        if v['online'] == 0 and v['absent'] == 0:
                                temp = temp + k + '\n'
                newItem = QTableWidgetItem(self.tr(temp[:-1]))
                newItem.setTextAlignment(Qt.AlignCenter)
                self.resultTable.setItem(Col, 2, newItem)

                # 早退
                temp = ''
                for k, v in persons.items():
                        if v['offline'] == 0 and v['disappear'] == 0:
                                temp = temp + k + '\n'
                newItem = QTableWidgetItem(self.tr(temp[:-1]))
                newItem.setTextAlignment(Qt.AlignCenter)
                self.resultTable.setItem(Col, 3, newItem)

                # 消失
                temp = ''
                for k, v in persons.items():
                        if v['offline'] == 0 and v['disappear'] == 1:
                                temp = temp + k + '\n'
                newItem = QTableWidgetItem(self.tr(temp[:-1]))
                newItem.setTextAlignment(Qt.AlignCenter)
                self.resultTable.setItem(Col, 4, newItem)

                Col = Col + 1
                self.resultTable.resizeRowsToContents()  

        temp = ''
        for key in persons:
                personDay = {}
                for r in xrange(2, R):
                        if sheet.cell(row=r, column=3).value == key:
                                day = sheet.cell(row=r, column=4).value
                                if not personDay.has_key(day):
                                        personDay[str(day)] = 1
                if float(len(days)) / float(len(personDay)) > 1.5:
                        temp = temp + key + ', '

        self.otherLineText.setText(self.tr(temp[:-2]))

class Comparison(QWidget):  
    def __init__(self, parent=None):  
        super(Comparison, self).__init__(parent)  
        file1Label = QPushButton(self.tr("选择扫描表格"))  
        file2Label = QPushButton(self.tr("选择参照表格"))
        self.file1LineText = QLineEdit() 
        self.file2LineText = QLineEdit()  
        key1Label = QLabel(self.tr("主键所在列"))
        key2Label = QLabel(self.tr("主键所在列"))
        self.key1LineText = QLineEdit("7")
        self.key2LineText = QLineEdit("7")
        title1Label = QLabel(self.tr("列名所在行"))
        title2Label = QLabel(self.tr("列名所在行"))
        self.title1LineText = QLineEdit("2")
        self.title2LineText = QLineEdit("2")
        load1PushButton = QPushButton(self.tr("载入扫描表格"))
        load2PushButton = QPushButton(self.tr("载入参考表格"))

        self.keyC1 = -1
        self.keyC2 = -1
        self.headR1 = -1
        self.headR2 = -1
        self.file1 = ''
        self.file2 = ''
        self.C1 = 0
        self.R1 = 0
        self.C2 = 0
        self.R2 = 0
        self.VC1 = 0
        self.VC2 = 0

        self.mappingTable = QTableWidget()
        self.mappingTable.setColumnCount(6)
        self.mappingTable.setRowCount(0)
        self.mappingTable.setHorizontalHeaderLabels([self.tr('扫描表列号'), self.tr('扫描表列名'), self.tr('扫描表列映射'), self.tr('参照表列映射'), self.tr('参照表列名'), self.tr('参照表列号')])
        self.mappingTable.verticalHeader().setVisible(False)

        noticeLable = QLabel(self.tr("请确认两个表格的列映射是否正确！如果存在差异请人工调整！"))
        startCompare = QPushButton(self.tr("开始对比"))
        self.resultTable = QTableWidget() 
        self.resultTable.setColumnCount(7)
        self.resultTable.setRowCount(0)
        self.resultTable.setHorizontalHeaderLabels([self.tr('不同处序号'), self.tr('所在行'), self.tr('所在列'), self.tr('主键'), self.tr('列名'), self.tr('扫描值'), self.tr('参考值')])
        self.resultTable.verticalHeader().setVisible(False)
        self.resultTable.setEditTriggers(QAbstractItemView.NoEditTriggers)

        freshLabel = QLabel(self.tr("新进人员"))
        self.freshLineText = QLineEdit()

        leftLayout = QGridLayout()
        leftLayout.addWidget(file1Label, 0, 0)
        leftLayout.addWidget(self.file1LineText, 0, 1)
        leftLayout.addWidget(key1Label, 1, 0)
        leftLayout.addWidget(self.key1LineText, 1, 1)
        leftLayout.addWidget(title1Label, 2, 0)
        leftLayout.addWidget(self.title1LineText, 2, 1)
        leftLayout.addWidget(load1PushButton, 3, 0)

        rightLayout = QGridLayout()
        rightLayout.addWidget(file2Label, 0, 0)
        rightLayout.addWidget(self.file2LineText, 0, 1)
        rightLayout.addWidget(key2Label, 1, 0)
        rightLayout.addWidget(self.key2LineText, 1, 1)
        rightLayout.addWidget(title2Label, 2, 0)
        rightLayout.addWidget(self.title2LineText, 2, 1)
        rightLayout.addWidget(load2PushButton, 3, 0)

        resultLayout = QGridLayout()
        resultLayout.addWidget(self.mappingTable, 0, 0, 1, 2)
        resultLayout.addWidget(noticeLable, 1, 0)
        resultLayout.addWidget(startCompare, 1, 1)
        resultLayout.addWidget(self.resultTable, 2, 0, 1, 2)

        layout = QGridLayout(self)
        layout.addLayout(leftLayout, 0, 0)
        layout.addLayout(rightLayout, 0, 1)
        layout.addLayout(resultLayout, 1, 0, 1, 2)
        layout.addWidget(freshLabel, 2, 0)
        layout.addWidget(self.freshLineText, 2, 1)

        self.connect(file1Label, SIGNAL("clicked()"), self.openFile1)
        self.connect(file2Label, SIGNAL("clicked()"), self.openFile2)
        self.connect(load1PushButton, SIGNAL("clicked()"), self.loadFile1)
        self.connect(load2PushButton, SIGNAL("clicked()"), self.loadFile2)
        self.connect(startCompare, SIGNAL("clicked()"), self.comparing)

    def openFile1(self):
        s = QFileDialog.getOpenFileName(self, "Please Choose A File", ".", "Xlsx files(*.xlsx)")
        self.file1LineText.setText(s)  
        s = str(s.toUtf8())
        self.file1 = s.split('/')[-1].decode("utf8")

    def openFile2(self):
        s = QFileDialog.getOpenFileName(self, "Please Choose A File", ".", "Xlsx files(*.xlsx)")
        self.file2LineText.setText(s)  
        s = str(s.toUtf8())
        self.file2 = s.split('/')[-1].decode("utf8")

    def loadFile1(self):
        self.keyC1 = self.key1LineText.text()
        self.headR1 = self.title1LineText.text()

        if self.keyC1 == '' or self.headR1 == '' or self.file1 == '':
                QMessageBox.critical(self, "错误", self.tr("请确认扫描表格、主键列、列名行已经选择！"))  
        else:
                work_book = openpyxl.load_workbook(filename=self.file1, data_only=True)
                self.file1 = work_book.get_sheet_by_name(work_book.get_sheet_names()[0])

                self.C1 = self.file1.get_highest_column()
                self.R1 = self.file1.get_highest_row()

                self.mappingTable.setRowCount(max(self.C1, self.C2))

                Col = 0

                for c in xrange(1, self.C1 + 1):
                        if self.file1.cell(row=int(self.headR1), column=c).value == None:
                                   continue
                        newItem = QTableWidgetItem(self.tr(str(c)))
                        newItem.setTextAlignment(Qt.AlignCenter)
                        self.mappingTable.setItem(Col, 0, newItem)
                        newItem = QTableWidgetItem(self.tr(str(c)))
                        newItem.setTextAlignment(Qt.AlignCenter)
                        self.mappingTable.setItem(Col, 2, newItem)
                        newItem = QTableWidgetItem(self.tr(self.file1.cell(row=int(self.headR1), column=c).value))
                        newItem.setTextAlignment(Qt.AlignCenter)
                        self.mappingTable.setItem(Col, 1, newItem)
                        Col = Col + 1
                # self.mappingTable.resizeColumnsToContents()

                self.VC1 = Col
                self.mappingTable.setRowCount(max(self.VC1, self.VC2))

    def loadFile2(self):
        self.keyC2 = self.key2LineText.text()
        self.headR2 = self.title2LineText.text()

        if self.keyC2 == '' or self.headR2 == '' or self.file2 == '':
                QMessageBox.critical(self, "错误", self.tr("请确认参照表格、主键列、列名行已经选择！"))  
        else:
                work_book = openpyxl.load_workbook(self.file2, data_only=True)
                self.file2 = work_book.get_sheet_by_name(work_book.get_sheet_names()[0])

                self.C2 = self.file2.get_highest_column()
                self.R2 = self.file2.get_highest_row()

                self.mappingTable.setRowCount(max(self.C1, self.C2))

                Col = 0

                for c in xrange(1, self.C2 + 1):
                        if self.file1.cell(row=int(self.headR1), column=c).value == None:
                                   continue
                        newItem = QTableWidgetItem(self.tr(str(c)))
                        newItem.setTextAlignment(Qt.AlignCenter)
                        self.mappingTable.setItem(Col, 3, newItem)
                        newItem = QTableWidgetItem(self.tr(str(c)))
                        newItem.setTextAlignment(Qt.AlignCenter)
                        self.mappingTable.setItem(Col, 5, newItem)
                        newItem = QTableWidgetItem(self.tr(self.file2.cell(row=int(self.headR2), column=c).value))
                        newItem.setTextAlignment(Qt.AlignCenter)
                        self.mappingTable.setItem(Col, 4, newItem)
                        Col = Col + 1

                self.VC2 = Col
                self.mappingTable.setRowCount(max(self.VC1, self.VC2))
          
    def comparing(self):
        tempC = max(self.VC1, self.VC2)
        mapping = {}
        for x in xrange(0, tempC):
                if self.mappingTable.item(x, 2).text() == None:
                    continue
                mapping[str(self.mappingTable.item(x, 2).text())] = str(self.mappingTable.item(x, 3).text())

        errorCount = 0
        freshmen = ''
        for r in xrange(int(self.headR1) + 1, self.R1 + 1):
                pKey = self.file1.cell(row=r, column=int(self.keyC1)).value
                if pKey == None:
                    continue
                fresh = True
                for r2 in xrange(int(self.headR2) + 1, self.R2 + 1):
                        if pKey == self.file2.cell(row=r2, column=int(self.keyC2)).value:
                                # 开始对比
                                fresh = False
                                for c in xrange(1, self.C1 + 1):
                                    if self.file1.cell(row=self.headR1, column=c).value == None:
                                        continue
                                    value1 = self.file1.cell(row=r, column=c).value
                                    value2 = self.file2.cell(row=r2, column=int(mapping[str(c)])).value

                                    if type(value1) == datetime.datetime:
                                        value1 = value1.strftime('%Y-%m-%d')
                                    elif not type(value1) == unicode:
                                        value1 = str(value1)

                                    if type(value2) == datetime.datetime:
                                        value2 = value2.strftime('%Y-%m-%d')
                                    elif not type(value2) == unicode:
                                        value2 = str(value2)

                                    if not value1 == value2:
                                                if value1 == 'None':
                                                    value1 = str('')
                                                if value2 == 'None':
                                                    value2 = str('')

                                                if len(value1) == 0 and len(value2) == 0:
                                                    continue

                                                print value1, value2
                                                print type(value1), type(value2)

                                                errorCount = errorCount + 1

                                                letter1 = ['','A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
                                                letter2 = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
                                                a = (c-1) / 26
                                                b = (c-1) % 26 
                                                column = letter1[a] + letter2[b]

                                                self.resultTable.setRowCount(errorCount)
                                                
                                                newItem = QTableWidgetItem(self.tr(str(errorCount)))
                                                newItem.setTextAlignment(Qt.AlignCenter)
                                                self.resultTable.setItem(errorCount - 1, 0, newItem)

                                                newItem = QTableWidgetItem(self.tr(str(r)))
                                                newItem.setTextAlignment(Qt.AlignCenter)
                                                self.resultTable.setItem(errorCount - 1, 1, newItem)

                                                newItem = QTableWidgetItem(self.tr(column))
                                                newItem.setTextAlignment(Qt.AlignCenter)
                                                self.resultTable.setItem(errorCount - 1, 2, newItem)

                                                newItem = QTableWidgetItem(self.tr(str(self.file1.cell(row=r, column=int(self.keyC1)).value)))
                                                newItem.setTextAlignment(Qt.AlignCenter)
                                                self.resultTable.setItem(errorCount - 1, 3, newItem)

                                                newItem = QTableWidgetItem(self.tr(self.file1.cell(row=int(self.headR1), column=c).value))
                                                newItem.setTextAlignment(Qt.AlignCenter)
                                                self.resultTable.setItem(errorCount - 1, 4, newItem)

                                                newItem = QTableWidgetItem(self.tr(value1))
                                                newItem.setTextAlignment(Qt.AlignCenter)
                                                self.resultTable.setItem(errorCount - 1, 5, newItem)

                                                newItem = QTableWidgetItem(self.tr(value2))
                                                newItem.setTextAlignment(Qt.AlignCenter)
                                                self.resultTable.setItem(errorCount - 1, 6, newItem)
                if fresh:
                    freshmen = freshmen + self.file1.cell(row=r, column=6).value + ';'

        self.freshLineText.setText(self.tr(freshmen[:-1]))

app = QApplication(sys.argv)  
main = XlsProcess()  
main.show()  
app.exec_()  
